import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(file_path, updated_file_path):
    """Process the Excel file by applying a discount and adding a chart."""
    wb = xl.load_workbook(file_path)
    sheet = wb.active  # Use the active sheet

    # Ensure the sheet has data before processing
    if sheet.max_row < 2 or sheet.max_column < 3:
        print(f"Skipping {file_path} (Not enough data)")
        return

    # Apply discount and save in column 4
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9  # Apply 10% discount
        sheet.cell(row, 4, corrected_price)  # Set value
        sheet.cell(1, 4, "Corrected Price")

    # Add a bar chart
    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    chart = BarChart()
    chart.add_data(values, titles_from_data=False)
    
    # Set title and axis labels
    chart.title = "Corrected Prices"
    chart.legend = None

    # Add the chart to the sheet
    sheet.add_chart(chart, "F5")

    # Save in the new folder
    wb.save(updated_file_path)
    print(f"Updated file saved: {updated_file_path}")
